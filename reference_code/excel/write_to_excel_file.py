import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers, PatternFill, Border, Side


def add_text_cell(work_sheet, column, row, value):
    # excel gives #NAME? error when editing cell with text like: --banana
    cell = work_sheet.cell(
        column=column,
        row=row,
        value=value
    )
    cell.number_format = numbers.FORMAT_TEXT
    return cell


def save(work_book, dir_name, file_name):
    file_name = os.path.join(dir_name, file_name)
    work_book.save(file_name)
    print('Saved exported data in "%s"' % file_name)


def run():
    dir_name = '/home/zhulien/work/data/excel'
    file_name = 'test.xlsx'

    work_book = Workbook()
    work_sheet = work_book.active
    color = 'FFFFE6'
    pf = PatternFill(fill_type='solid', start_color=color, end_color=color)
    thin = Side(style='thin', color='000000')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    min_col, max_col = 1, 4
    min_row, max_row = 1, 4

    for col in range(min_col, max_col + 1):
        column_letter = get_column_letter(col)
        work_sheet.column_dimensions[column_letter].width = 20

    for col in range(min_col, max_col + 1):
        for row in range(min_row, max_row + 1):
            cell = add_text_cell(work_sheet, col, row, 'HELLO BANANA %s-%s' % (row, col))
            cell.fill = pf
            cell.border = border

    work_sheet.title = 'Bananas'
    # below if for adding filters to all columns if the first row is with column names
    #work_sheet.auto_filter.ref = work_sheet.dimensions
    
    work_sheet = work_book.create_sheet('Mangos')
    add_text_cell(work_sheet, 1, 1, 'One mango')
    
    
    
    save(work_book, dir_name, file_name)


if __name__ == '__main__':
    run()
